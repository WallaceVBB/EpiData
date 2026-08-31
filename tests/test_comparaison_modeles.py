"""
Script de benchmark : compare LinearSVC (avec/sans calibration) et SGDClassifier
pour la tâche de classification `base_variante`, sur plusieurs filtres de fréquence
minimale de classe.

Objectif : mesurer, pour chaque combinaison (modèle x filtre de classe) :
  - le temps d'entraînement
  - l'accuracy sur un jeu de test tenu à l'écart (jamais vu à l'entraînement)
  - le F1-score macro (traite toutes les classes également, car
    beaucoup de classes ont peu d'exemples)
  - le F1-score pondéré (reflète mieux la performance "globale" perçue)
  - la confiance moyenne du modèle sur ses prédictions correctes vs incorrectes
  - Fiabilité par categorie (precision / rappel / F1 / support / confiance moyenne pour chaque classe de
    base_variante) et exporte le détail dans un fichier Excel

Résultats :
  - affichés dans la console (table rich)
  - résumé global sauvegardé en CSV dans tests/resultats_benchmark_<timestamp>.csv
  - détail par catégorie sauvegardé en Excel dans
    tests/fiabilite_par_categorie_<timestamp>.xlsx (un onglet par combinaison
    filtre de classe x modèle)
"""

import os
import sys
import time
import warnings
from dataclasses import dataclass
from datetime import datetime

import numpy as np
import pandas as pd
import sqlite3

from sklearn.exceptions import ConvergenceWarning
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import SGDClassifier
from sklearn.metrics import accuracy_score, f1_score, classification_report
from sklearn.model_selection import train_test_split
from sklearn.svm import LinearSVC
from sklearn.calibration import CalibratedClassifierCV

from rich.console import Console
from rich.table import Table

from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

# Permet d'importer utils.py / services.py depuis la racine du projet,
# même si ce script est lancé depuis le dossier tests/.
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from utils import BD_ENTRAINEMENT, nettoyer_texte  # noqa: E402

console = Console()

# On veut voir tous les warnings de convergence/calibration pour évaluer
# honnêtement, mais sans qu'ils polluent trop l'affichage rich.
warnings.filterwarnings("ignore", category=ConvergenceWarning)


# ---------------------------------------------------------------------------
# Configuration du benchmark
# ---------------------------------------------------------------------------

# Filtres de fréquence minimale de classe à tester (donnees['base_variante']
# doit apparaître au moins ce nombre de fois pour être conservée).
FILTRES_CLASSE = [15, 20]

# Proportion des données réservée à l'évaluation finale (jamais utilisée pour
# l'entraînement d'aucun modèle, ni pour la calibration).
TEST_SIZE = 0.2

# Graine fixe pour la reproductibilité des comparaisons.
RANDOM_STATE = 42

# Paramètres TF-IDF (identiques à ceux du code de production, pour que la
# comparaison porte uniquement sur le classifieur).
TFIDF_PARAMS = dict(max_features=3000, ngram_range=(1, 2), sublinear_tf=True, min_df=3)


@dataclass
class ResultatBenchmark:
    filtre_classe: int
    nb_lignes: int
    nb_classes: int
    modele: str
    temps_entrainement_s: float
    accuracy: float
    f1_macro: float
    f1_pondere: float
    confiance_moyenne_correct: float
    confiance_moyenne_incorrect: float
    notes: str = ""


def charger_donnees(bd_path):
    """Charge les colonnes designation/base_variante depuis la base d'entraînement."""
    with sqlite3.connect(bd_path) as conn:
        donnees = pd.read_sql_query("SELECT designation, base_variante FROM entrainement", conn)
    return donnees


def filtrer_par_classe(donnees, seuil_min):
    """Ne garde que les lignes dont la classe apparaît au moins `seuil_min` fois."""
    compte = donnees["base_variante"].value_counts()
    return donnees[donnees["base_variante"].map(compte) >= seuil_min].reset_index(drop=True)


def preparer_split(donnees, test_size, random_state):
    """Sépare un jeu de test tenu à l'écart, stratifié par classe.

    Ce split est commun à tous les modèles testés pour ce filtre de classe :
    c'est ce qui rend la comparaison équitable (même train, même test).
    """
    textes = donnees["designation"].apply(nettoyer_texte).tolist()
    labels = donnees["base_variante"].tolist()
    return train_test_split(
        textes, labels, test_size=test_size, stratify=labels, random_state=random_state
    )


def confiance_correct_incorrect(y_test, predictions, probabilites, classes_):
    """Calcule la confiance moyenne (proba max) séparément pour les prédictions
    correctes et incorrectes. Un bon modèle calibré doit avoir une confiance
    nettement plus élevée sur ses prédictions correctes."""
    proba_max = probabilites.max(axis=1)
    est_correct = np.array(predictions) == np.array(y_test)

    conf_correct = proba_max[est_correct].mean() if est_correct.any() else float("nan")
    conf_incorrect = proba_max[~est_correct].mean() if (~est_correct).any() else float("nan")
    return conf_correct, conf_incorrect


def evaluer(y_test, predictions, probabilites, classes_):
    accuracy = accuracy_score(y_test, predictions)
    f1_macro = f1_score(y_test, predictions, average="macro", zero_division=0)
    f1_pondere = f1_score(y_test, predictions, average="weighted", zero_division=0)
    conf_correct, conf_incorrect = confiance_correct_incorrect(y_test, predictions, probabilites, classes_)
    return accuracy, f1_macro, f1_pondere, conf_correct, conf_incorrect


def evaluer_fiabilite_par_classe(y_test, predictions, probabilites, classes_, filtre_classe, modele_nom):
    """Calcule, pour CHAQUE classe de base_variante, la précision/rappel/F1/support
    (via classification_report) ainsi que la confiance moyenne du modèle sur les
    exemples réellement de cette classe (globale, correcte, incorrecte).

    Retourne un DataFrame avec une ligne par classe.
    """
    y_test_arr = np.array(y_test)
    predictions_arr = np.array(predictions)
    proba_max = probabilites.max(axis=1)

    rapport = classification_report(
        y_test_arr, predictions_arr, output_dict=True, zero_division=0
    )

    lignes = []
    for classe in classes_:
        classe_str = str(classe)
        stats = rapport.get(classe_str, {})

        masque_classe = y_test_arr == classe
        nb_exemples = int(masque_classe.sum())

        if nb_exemples == 0:
            confiance_moyenne = conf_correct = conf_incorrect = float("nan")
        else:
            est_correct = predictions_arr[masque_classe] == y_test_arr[masque_classe]
            confiances_classe = proba_max[masque_classe]
            confiance_moyenne = confiances_classe.mean()
            conf_correct = confiances_classe[est_correct].mean() if est_correct.any() else float("nan")
            conf_incorrect = confiances_classe[~est_correct].mean() if (~est_correct).any() else float("nan")

        lignes.append(
            {
                "filtre_classe": filtre_classe,
                "modele": modele_nom,
                "categorie": classe,
                "support_test": nb_exemples,
                "precision": stats.get("precision", float("nan")),
                "rappel": stats.get("recall", float("nan")),
                "f1": stats.get("f1-score", float("nan")),
                "confiance_moyenne": confiance_moyenne,
                "confiance_moyenne_si_correct": conf_correct,
                "confiance_moyenne_si_incorrect": conf_incorrect,
            }
        )

    df = pd.DataFrame(lignes).sort_values("f1", ascending=True, na_position="first")
    return df


def entrainer_linearsvc_calibre(X_train_texts, y_train, X_test_texts, vectoriseur):
    """LinearSVC (dual=False) + calibration via CalibratedClassifierCV(cv=3).

    C'est l'équivalent du code de production actuel (sans le split de calibration
    séparé, pour rester sur une base standard et reproductible ici) : mesure le
    coût réel de la calibration multi-fold.
    """
    X_train = vectoriseur.transform(X_train_texts)
    X_test = vectoriseur.transform(X_test_texts)

    debut = time.perf_counter()
    modele = CalibratedClassifierCV(
        LinearSVC(class_weight="balanced", max_iter=1000, dual=False, random_state=RANDOM_STATE),
        cv=3,
    ).fit(X_train, y_train)
    temps = time.perf_counter() - debut

    predictions = modele.predict(X_test)
    probabilites = modele.predict_proba(X_test)
    return modele, temps, predictions, probabilites


def entrainer_sgd(X_train_texts, y_train, X_test_texts, vectoriseur):
    """SGDClassifier avec loss='log_loss' : probabilites natives, pas de calibration
    séparée nécessaire."""
    X_train = vectoriseur.transform(X_train_texts)
    X_test = vectoriseur.transform(X_test_texts)

    debut = time.perf_counter()
    modele = SGDClassifier(
        loss="log_loss",
        class_weight="balanced",
        max_iter=1000,
        n_jobs=-1,
        random_state=RANDOM_STATE,
    ).fit(X_train, y_train)
    temps = time.perf_counter() - debut

    predictions = modele.predict(X_test)
    probabilites = modele.predict_proba(X_test)
    return modele, temps, predictions, probabilites


def executer_benchmark():
    console.print(f"[bold cyan]Chargement de la base d'entraînement : {BD_ENTRAINEMENT}")
    donnees_completes = charger_donnees(BD_ENTRAINEMENT)
    console.print(f"[cyan]{len(donnees_completes)} lignes chargées au total.\n")

    resultats = []
    fiabilite_par_classe = []  # liste de DataFrames, un par combinaison (filtre, modele)

    for seuil in FILTRES_CLASSE:
        console.print(f"[bold yellow]--- Filtre : classes avec >= {seuil} occurrences ---")
        donnees = filtrer_par_classe(donnees_completes, seuil)
        nb_classes = donnees["base_variante"].nunique()
        console.print(f"[yellow]{len(donnees)} lignes, {nb_classes} classes retenues.")

        if nb_classes < 2 or len(donnees) < 20:
            console.print("[red]Pas assez de données pour ce filtre, on saute.\n")
            continue

        X_train_texts, X_test_texts, y_train, y_test = preparer_split(
            donnees, TEST_SIZE, RANDOM_STATE
        )

        # Le vectoriseur TF-IDF est entraîné une seule fois sur X_train, partagé
        # entre les deux modèles pour isoler la comparaison au classifieur seul.
        vectoriseur = TfidfVectorizer(**TFIDF_PARAMS)
        vectoriseur.fit(X_train_texts)

        # --- LinearSVC calibré ---
        try:
            modele_svc, temps_svc, pred_svc, proba_svc = entrainer_linearsvc_calibre(
                X_train_texts, y_train, X_test_texts, vectoriseur
            )
            acc, f1m, f1p, conf_c, conf_i = evaluer(y_test, pred_svc, proba_svc, modele_svc.classes_)
            resultats.append(
                ResultatBenchmark(
                    filtre_classe=seuil,
                    nb_lignes=len(donnees),
                    nb_classes=nb_classes,
                    modele="LinearSVC+Calibration",
                    temps_entrainement_s=temps_svc,
                    accuracy=acc,
                    f1_macro=f1m,
                    f1_pondere=f1p,
                    confiance_moyenne_correct=conf_c,
                    confiance_moyenne_incorrect=conf_i,
                )
            )
            fiabilite_par_classe.append(
                evaluer_fiabilite_par_classe(
                    y_test, pred_svc, proba_svc, modele_svc.classes_, seuil, "LinearSVC+Calibration"
                )
            )
            console.print(f"[green]LinearSVC+Calibration : {temps_svc:.1f}s, accuracy={acc:.3f}")
        except Exception as e:
            console.print(f"[red]Erreur LinearSVC : {e}")

        # --- SGDClassifier ---
        try:
            modele_sgd, temps_sgd, pred_sgd, proba_sgd = entrainer_sgd(
                X_train_texts, y_train, X_test_texts, vectoriseur
            )
            acc, f1m, f1p, conf_c, conf_i = evaluer(y_test, pred_sgd, proba_sgd, modele_sgd.classes_)
            resultats.append(
                ResultatBenchmark(
                    filtre_classe=seuil,
                    nb_lignes=len(donnees),
                    nb_classes=nb_classes,
                    modele="SGDClassifier",
                    temps_entrainement_s=temps_sgd,
                    accuracy=acc,
                    f1_macro=f1m,
                    f1_pondere=f1p,
                    confiance_moyenne_correct=conf_c,
                    confiance_moyenne_incorrect=conf_i,
                )
            )
            fiabilite_par_classe.append(
                evaluer_fiabilite_par_classe(
                    y_test, pred_sgd, proba_sgd, modele_sgd.classes_, seuil, "SGDClassifier"
                )
            )
            console.print(f"[green]SGDClassifier : {temps_sgd:.1f}s, accuracy={acc:.3f}\n")
        except Exception as e:
            console.print(f"[red]Erreur SGDClassifier : {e}\n")

    return resultats, fiabilite_par_classe


def afficher_resultats(resultats):
    table = Table(title="Comparaison LinearSVC vs SGDClassifier")
    table.add_column("Filtre classe", justify="right")
    table.add_column("Lignes", justify="right")
    table.add_column("Classes", justify="right")
    table.add_column("Modèle")
    table.add_column("Temps (s)", justify="right")
    table.add_column("Accuracy", justify="right")
    table.add_column("F1 macro", justify="right")
    table.add_column("F1 pondéré", justify="right")
    table.add_column("Conf. correct", justify="right")
    table.add_column("Conf. incorrect", justify="right")

    for r in resultats:
        table.add_row(
            str(r.filtre_classe),
            str(r.nb_lignes),
            str(r.nb_classes),
            r.modele,
            f"{r.temps_entrainement_s:.1f}",
            f"{r.accuracy:.3f}",
            f"{r.f1_macro:.3f}",
            f"{r.f1_pondere:.3f}",
            f"{r.confiance_moyenne_correct:.3f}",
            f"{r.confiance_moyenne_incorrect:.3f}",
        )

    console.print(table)


def sauvegarder_csv(resultats):
    df = pd.DataFrame([r.__dict__ for r in resultats])
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    chemin = os.path.join(os.path.dirname(__file__), f"resultats_benchmark_{timestamp}.csv")
    df.to_csv(chemin, index=False)
    console.print(f"\n[bold cyan]Résultats sauvegardés dans : {chemin}")


def _nom_onglet(filtre_classe, modele):
    """Les noms d'onglet Excel sont limités à 31 caractères."""
    nom = f"f{filtre_classe}_{modele}"
    return nom[:31]


def sauvegarder_excel_fiabilite(fiabilite_par_classe):
    """Exporte le détail de fiabilité par catégorie en Excel : un onglet par
    combinaison (filtre de classe, modèle), trié du F1 le plus faible au plus
    élevé (les catégories problématiques en haut), avec une échelle de couleur
    rouge -> vert sur la colonne F1 pour un repérage visuel immédiat.
    """
    if not fiabilite_par_classe:
        console.print("[yellow]Aucune donnée de fiabilité par catégorie à exporter.")
        return

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    chemin = os.path.join(os.path.dirname(__file__), f"fiabilite_par_categorie_{timestamp}.xlsx")

    with pd.ExcelWriter(chemin, engine="openpyxl") as writer:
        for df in fiabilite_par_classe:
            filtre_classe = df["filtre_classe"].iloc[0]
            modele = df["modele"].iloc[0]
            colonnes_export = [
                "categorie",
                "support_test",
                "precision",
                "rappel",
                "f1",
                "confiance_moyenne",
                "confiance_moyenne_si_correct",
                "confiance_moyenne_si_incorrect",
            ]
            df_export = df[colonnes_export].round(3)
            nom_onglet = _nom_onglet(filtre_classe, modele)
            df_export.to_excel(writer, sheet_name=nom_onglet, index=False)

            feuille = writer.sheets[nom_onglet]

            # Largeur de colonnes raisonnable
            for i, col in enumerate(colonnes_export, start=1):
                largeur = max(12, len(col) + 2)
                feuille.column_dimensions[get_column_letter(i)].width = largeur

            # Échelle de couleur rouge -> jaune -> vert sur la colonne F1
            idx_f1 = colonnes_export.index("f1") + 1
            lettre_f1 = get_column_letter(idx_f1)
            nb_lignes = len(df_export)
            if nb_lignes > 0:
                plage = f"{lettre_f1}2:{lettre_f1}{nb_lignes + 1}"
                regle = ColorScaleRule(
                    start_type="num", start_value=0, start_color="F8696B",     # rouge
                    mid_type="num", mid_value=0.5, mid_color="FFEB84",         # jaune
                    end_type="num", end_value=1, end_color="63BE7B",           # vert
                )
                feuille.conditional_formatting.add(plage, regle)

            feuille.freeze_panes = "A2"

    console.print(f"[bold cyan]Fiabilité par catégorie sauvegardée dans : {chemin}")
    return chemin


if __name__ == "__main__":
    resultats, fiabilite_par_classe = executer_benchmark()
    if resultats:
        afficher_resultats(resultats)
        sauvegarder_csv(resultats)
        sauvegarder_excel_fiabilite(fiabilite_par_classe)
    else:
        console.print("[bold red]Aucun résultat produit — vérifiez la base d'entraînement.")